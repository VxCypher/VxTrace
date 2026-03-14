from __future__ import annotations

from pathlib import Path
from typing import Any
from PIL import Image, ExifTags
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


def image_metadata(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {}
    with Image.open(path) as img:
        result["width"] = img.width
        result["height"] = img.height
        result["mode"] = img.mode
        result["format"] = img.format
        result["frames"] = getattr(img, "n_frames", 1)

        exif_out: dict[str, Any] = {}
        try:
            exif = img.getexif()
            if exif:
                tag_map = {v: k for k, v in ExifTags.TAGS.items()}
                for tag_id, value in exif.items():
                    name = ExifTags.TAGS.get(tag_id, str(tag_id))
                    if isinstance(value, bytes):
                        continue
                    exif_out[name] = str(value)
        except Exception as exc:
            exif_out["_error"] = f"EXIF read failed: {exc}"
        if exif_out:
            result["exif"] = exif_out
    return result


def pdf_metadata(path: Path) -> dict[str, Any]:
    reader = PdfReader(str(path))
    meta = reader.metadata or {}
    out = {
        "pages": len(reader.pages),
        "encrypted": reader.is_encrypted,
        "pdf": {},
    }
    for key, value in meta.items():
        clean_key = str(key).lstrip("/")
        out["pdf"][clean_key] = None if value is None else str(value)
    return out


def docx_metadata(path: Path) -> dict[str, Any]:
    doc = Document(str(path))
    props = doc.core_properties
    fields = [
        "author", "category", "comments", "content_status", "created",
        "identifier", "keywords", "language", "last_modified_by", "last_printed",
        "modified", "revision", "subject", "title", "version",
    ]
    core = {}
    for field in fields:
        core[field] = str(getattr(props, field))
    return {"docx": core}


def xlsx_metadata(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=str(path), read_only=True)
    props = wb.properties
    fields = [
        "creator", "title", "description", "subject", "identifier", "language",
        "lastModifiedBy", "category", "contentStatus", "revision", "version",
        "created", "modified",
    ]
    core = {}
    for field in fields:
        core[field] = str(getattr(props, field, None))
    return {"xlsx": core, "worksheets": list(wb.sheetnames)}
